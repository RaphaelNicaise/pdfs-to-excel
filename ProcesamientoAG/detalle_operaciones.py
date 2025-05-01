import os
import shutil
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import xlwings
from xlwings import Book


def get_asset_path(relative_path):
    """Obtiene la ruta correcta del archivo, ya sea en desarrollo o en el ejecutable."""
    try:
        # Cuando se ejecuta como ejecutable empaquetado
        base_path = sys._MEIPASS
    except AttributeError:
        # Cuando se ejecuta como script en desarrollo
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def main_process_detalle_operacion(ruta_excel_base, ruta_destino, excel_datos, anio, mes):
    """
    Genera un Excel con detalle de operaciones a partir de una plantilla.

    Args:
        ruta_excel_base (str): Ruta de la plantilla base (siempre será 'assets/plantilla_detalle_operaciones.xlsx').
        ruta_destino (str): Carpeta donde se guardará el archivo resultante.
        excel_datos (str): Ruta del archivo Excel con los datos.
        anio (int): Año a insertar.
        mes (str): Mes a insertar.
    """
    ruta_excel_base = get_asset_path(ruta_excel_base)

    # Crear nombre del nuevo archivo
    nombre_empresa = os.path.basename(excel_datos).split('AG-')[1].split('.xlsx')[0]
    nombre_archivo = "Detalle Proceso " + nombre_empresa + ".xlsx"
    ruta_final = os.path.join(ruta_destino, nombre_archivo)

    # Copiar plantilla a destino
    shutil.copy(ruta_excel_base, ruta_final)

    # Abrir archivo copiado
    wb = load_workbook(ruta_final)
    wb.template = False
    ws = wb.active

    # Insertar mes y año
    ws['B8'] = mes
    ws['B8'].alignment = Alignment(horizontal='left', vertical='center')

    ws['E8'] = anio
    ws['E8'].alignment = Alignment(horizontal='left', vertical='center')

    # Nombre de la firma (harcodeado por ahora)
    
    ws['B10'] = nombre_empresa
    ws['B10'].alignment = Alignment(horizontal='left', vertical='center')

    # Leer datos desde el archivo Excel de datos
    df = pd.read_excel(excel_datos, engine='openpyxl')

    df_final = df.loc[:, ['FECHA CARGA', 'MIC - DTA', 'D.D.T', 'ORDEN', 'FACTURA Nº']]
    df_final['FECHA CARGA'] = pd.to_datetime(df_final['FECHA CARGA'], dayfirst=True, errors='coerce')
    # Ordenar por la columna 'FECHA CARGA' de menor a mayor
    df_final = df_final.sort_values(by='FECHA CARGA', ascending=True).reset_index(drop=True)
    df_final['FECHA CARGA'] = df_final['FECHA CARGA'].dt.strftime('%d/%m/%Y')
    
    # Insertar total de unidades
    total_unidades = len(df_final)
    ws['B11'] = total_unidades
    ws['B11'].alignment = Alignment(horizontal='left', vertical='center')

    # Insertar datos desde fila 14 (índice 13)
    for i, row in df_final.iterrows():
        for j, campo in enumerate(['FECHA CARGA', 'MIC - DTA', 'D.D.T', 'ORDEN', 'FACTURA Nº']):
            ws.cell(row=14 + i, column=1 + j, value=row[campo])

    # Guardar
    wb.save(ruta_final)
    print(f"Archivo guardado en: {ruta_final}")
    
    wb = Book(ruta_final)
    ws =  wb.sheets[0]
    
    
    
    img = get_asset_path('assets/imagen.jpg')
    ws.pictures.add(
        img,
        top=ws.range('A1').top,
        left=ws.range('A1').left,
        width=ws.range('E5').left + ws.range('E5').width - ws.range('A1').left,
        height=ws.range('E5').top + ws.range('E5').height - ws.range('A1').top
    )
    wb.save()
    wb.app.quit()
    print(f"Archivo guardado en: {ruta_final}")